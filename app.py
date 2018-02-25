import requests
import shutil
import openpyxl
import xlrd
from openpyxl.workbook import Workbook as openpyxlWorkbook
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np

def download_file(url):
    split = url.split("/")
    file_name = split[len(split) - 1]
    print("Downloaded " + file_name)
    r = requests.get(url, stream=True)
    if r.status_code == 200:
        with open(file_name, 'wb') as f:
            r.raw.decode_content = True
            shutil.copyfileobj(r.raw, f)


download_file("https://ucr.fbi.gov/crime-in-the-u.s/2013/crime-in-the-u.s.-2013/tables/1tabledatadecoverviewpdf/"
              "table_1_crime_in_the_united_states_by_volume_and_rate_per_100000_inhabitants_1994-2013.xls/output.xls")
download_file("https://ucr.fbi.gov/crime-in-the-u.s/2013/crime-in-the-u.s.-2013/tables/table-8/table_8_offenses_known_to_law_enforcement_by_state_by_city_2013.xls")

def average_crime_per_year(sheet, row_num):
    return sheet.cell(row_num, 3).value + sheet.cell(row_num, 13).value


def answer(avg1995, avg2013):
    print(avg1995)
    print(avg2013)

    if avg1995 > avg2013:
        return "Crime decresed since 1995 until 2013 by " + str(round(100 - avg2013 * 100 / avg1995, 2)) + "%"
    else:
        return "Crime increased since 1995 until 2013 by " + str(round(100 - avg1995 * 100 / avg2013, 2)) + "%"


def first_question():
    xlsBook = xlrd.open_workbook('./output.xls')
    sheet = xlsBook.sheet_by_index(0)

    print(answer(average_crime_per_year(sheet, 4), average_crime_per_year(sheet, 23)))


first_question()

def average_murder_per_year(sheet, row_num):
    return sheet.cell(row_num, 5).value
def average_rape_per_year(sheet, row_num):
    return sheet.cell(row_num, 7).value
def average_robbery_per_year(sheet, row_num):
    return sheet.cell(row_num, 9).value
def average_aggravated_assault_per_year(sheet, row_num):
    return sheet.cell(row_num, 11).value
def average_burglary_per_year(sheet, row_num):
    return sheet.cell(row_num, 15).value
def average_larceny_per_year(sheet, row_num):
    return sheet.cell(row_num, 17).value
def average_vehicle_theft_per_year(sheet, row_num):
    return sheet.cell(row_num, 19).value

def second_question():
    xlsBook = xlrd.open_workbook('./output.xls')
    sheet = xlsBook.sheet_by_index(0)

    crimes_1995 = np.array([average_murder_per_year(sheet, 4), average_rape_per_year(sheet, 4), average_robbery_per_year(sheet, 4), average_aggravated_assault_per_year(sheet, 4), average_burglary_per_year(sheet, 4), average_larceny_per_year(sheet, 4), average_vehicle_theft_per_year(sheet, 4)])
    crimes_2013 = np.array([average_murder_per_year(sheet, 23), average_rape_per_year(sheet, 23), average_robbery_per_year(sheet, 23), average_aggravated_assault_per_year(sheet, 23), average_burglary_per_year(sheet, 23), average_larceny_per_year(sheet, 23), average_vehicle_theft_per_year(sheet, 23)])

    print(crimes_1995)
    print(crimes_2013)

    fig = plt.figure()
    x = np.array([0,1,2,3,4,5,6])
    my_xticks= ["Murder", "Rape", "Robbery", "Assault", "Burglary", "Larceny", "GTA"]
    plt.xticks(x, my_xticks)    
    _1995 = plt.bar(my_xticks, crimes_1995, color='r')
    _2013 = plt.bar(my_xticks, crimes_2013, color='b')
    plt.legend((_1995[0],_2013[0]),("1995", "2013"))
    plt.title("Crime Rates")
    plt.xlabel("Crime Type")
    plt.ylabel("Rate")

    fig.savefig('test.png')
    
second_question()